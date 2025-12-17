import openpyxl
from unidecode import unidecode

# --------------------------------------------------
# Função de limpeza sem alterar formatação
# --------------------------------------------------
def limpar_texto(valor):
    if not isinstance(valor, str):
        return valor
    
    texto = valor.lower().strip()
    texto = unidecode(texto)

    # padroniza (presencial) e (remoto)
    texto = texto.replace("(presencial)", "(presencial)")
    texto = texto.replace("(remoto)", "(remoto)")

    # remove espaços duplicados
    while "  " in texto:
        texto = texto.replace("  ", " ")

    return texto


# --------------------------------------------------
# Colunas que DEVEM ser limpas (somente essas!)
# --------------------------------------------------
COLUNAS_LIMPAR = ["produtos", "solucao", "solução"]


# --------------------------------------------------
# Função principal
# --------------------------------------------------
def limpar_planilha_sem_formatacao(caminho_excel):
    try:
        wb = openpyxl.load_workbook(caminho_excel)
    except FileNotFoundError:
        print(f"❌ Arquivo não encontrado: {caminho_excel}")
        return

    ws = wb.active

    # detectar quais colunas existem e devem ser limpas
    colunas_para_limpar = {}
    for col in range(1, ws.max_column + 1):
        nome_col = str(ws.cell(row=1, column=col).value).lower().strip()
        if nome_col in COLUNAS_LIMPAR:
            colunas_para_limpar[nome_col] = col

    if not colunas_para_limpar:
        print(f"⚠ Nenhuma coluna para limpar encontrada em {caminho_excel}")
    else:
        print(f"✔ Colunas que serão limpas em {caminho_excel}: {list(colunas_para_limpar.keys())}")

    # aplicar limpeza SOMENTE nas colunas detectadas
    for nome, col in colunas_para_limpar.items():
        for linha in range(2, ws.max_row + 1):
            valor = ws.cell(row=linha, column=col).value
            ws.cell(row=linha, column=col).value = limpar_texto(valor)

    wb.save(caminho_excel)
    print(f"✔ Conteúdo limpo (formatação preservada): {caminho_excel}")


# --------------------------------------------------
# Execução automática para os dois arquivos
# --------------------------------------------------
print("Iniciando limpeza com preservação de formatação...")

limpar_planilha_sem_formatacao("BASE_Interna_Conferida_2025.xlsx")
limpar_planilha_sem_formatacao("Planlilha_Consulta_Repasses_Credenciados_2025.xlsx")

print("🎉 Limpeza concluída sem alterar formatação!")
